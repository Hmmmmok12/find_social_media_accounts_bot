import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog

import requests
import json

def select_file():
    file_path = filedialog.askopenfilename(filetypes=(("All files", "*"),))
    file_entry.delete(0, tk.END)
    file_entry.insert(0, file_path)

def api_key():
    api = api_entry.get()
    return api

def find():
    try:
        # get Apollo API key
        api = api_key()

        # Load the input Excel file
        input_file_path = file_entry.get()

        # Open the input workbook and get the first sheet
        input_workbook = openpyxl.load_workbook(input_file_path)
        input_sheet = input_workbook.active

        # Create a new workbook for the output file
        output_workbook = Workbook()
        output_sheet = output_workbook.active

        output_sheet.append(["Company Name", "Name", "Title", "Headline", "Email", "LinkedIn", "Twitter", "Facebook"])

        # Loop through each row in the input sheet
        for row in input_sheet.iter_rows(min_row=2, values_only=True):  # Assuming the data starts from the 2nd row
            if any(cell_value is not None and cell_value != "" for cell_value in row):
                company_name = row[0] if row[0] else ""
                domain =  row[1] if row[1] else ""
                title = row[2] if row[2] else ""

                # Use apollo.io API
                url = "https://api.apollo.io/v1/mixed_people/search"

                data = {
                    "api_key": f"{api}",
                    "q_organization_domains": f"{domain}",
                    "page" : 1,
                    "person_titles" : [f"{title}"]
                }
        
                headers = {
                    'Cache-Control': 'no-cache',
                    'Content-Type': 'application/json'
                }
        
                response = requests.request("POST", url, headers=headers, json=data)

                res = json.loads(response.text)

                for i in range(len(res["people"])):
                    output_sheet.append([company_name,
                                         res["people"][i]["name"],
                                         res["people"][i]["title"],
                                         res["people"][i]["headline"],
                                         res["people"][i]["email"],
                                         res["people"][i]["linkedin_url"],
                                         res["people"][i]["twitter_url"],
                                         res["people"][i]["facebook_url"]])

        # Save the output workbook with the Twitter profiles for each row
        output_file_path = "output_file.xlsx"
        output_workbook.save(output_file_path)
        status_label.config(text=f"Output saved to: {output_file_path}")

    except Exception:
        status_label.config(text="Error occurred while processing the Excel file.")


root = tk.Tk()
root.title("Find Social Media Accounts")

# create the main frame
main_frame = tk.Frame(root, padx=10, pady=10)
main_frame.pack()

# input apollo API key
api_frame = tk.LabelFrame(main_frame, text="Input Apollo API key")
api_frame.pack(fill="x", padx=10, pady=10)

api_entry = tk.Entry(api_frame, width=40)
api_entry.pack(side="left", padx=10, pady=5)

api_botton = tk.Button(api_frame, text="Enter", command=api_key)
api_botton.pack(side="left", padx=10, pady=5)

# create the file frame
file_frame = tk.LabelFrame(main_frame, text="Upload Excel file")
file_frame.pack(fill="x", padx=10, pady=10)

file_entry = tk.Entry(file_frame, width=40)
file_entry.pack(side="left", padx=10, pady=5)

file_button = tk.Button(file_frame, text="Upload", command=select_file)
file_button.pack(side="left", padx=10, pady=5)

# create the status label
status_label = tk.Label(main_frame, text="", font=("Arial", 12))
status_label.pack(pady=10)

# create the send button
send_button = tk.Button(main_frame, text="Find Social Media Accounts", command=find)
send_button.pack(pady=5)

root.mainloop()